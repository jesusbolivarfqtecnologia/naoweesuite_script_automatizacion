import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def leer_cedula(ws_apu) -> Any:
	"""Lee la cédula desde la hoja 'APU' en la celda L6."""
	return ws_apu["L6"].value


def _to_str(v: Any) -> str:
	return "" if v is None else str(v)


def _categoria_from_id(id_val: Any) -> str:
	"""Obtiene el código de categoría a partir del id (parte antes del punto). Ej: '14.1' -> '14'."""
	if id_val is None:
		return ""
	s = str(id_val)
	# Normalizar notación de floats con coma si apareciera
	s = s.replace(",", ".")
	return s.split(".")[0]


def _sum_safe(values: List[Any]) -> float:
	total = 0.0
	for v in values:
		try:
			if v is None or (isinstance(v, str) and v.strip() == ""):
				continue
			total += float(v)
		except Exception:
			# Si no es numérico, lo ignoramos en la suma
			continue
	return total


def _round2_if_number(v: Any) -> Any:
	"""Convierte a float y redondea a 2 decimales si es numérico; de lo contrario devuelve el valor original."""
	try:
		if v is None or (isinstance(v, str) and v.strip() == ""):
			return v
		# Evitar convertir strings no numéricos como 'ABC'
		f = float(v)
		return round(f, 2)
	except Exception:
		return v


def _clean_discounts_in_details(details: List[Dict[str, Any]]) -> None:
	"""Filtra descuentos conservando solo aquellos con algún valor numérico positivo (> 0).
	Si no queda ninguno, elimina la clave 'discounts' del quantity_detail.
	"""
	NULL_KEYS = ["height", "width", "length", "area", "quantity", "subtotal"]

	def _has_positive_number(disc: Dict[str, Any]) -> bool:
		for k in NULL_KEYS:
			v = disc.get(k)
			try:
				if v is None or (isinstance(v, str) and v.strip() == ""):
					continue
				if float(v) > 0:
					return True
			except Exception:
				continue
		return False

	for d in details:
		discounts = d.get("discounts")
		if not discounts:
			d.pop("discounts", None)
			continue
		filtered = []
		for disc in discounts:
			if _has_positive_number(disc):
				filtered.append(disc)
		if filtered:
			d["discounts"] = filtered
		else:
			d.pop("discounts", None)


def _zero_nulls_in_details(details: List[Dict[str, Any]]) -> None:
	"""Convierte valores nulos en 0.0 para los campos numéricos de cada detail y sus discounts."""
	NUM_KEYS = ["height", "width", "length", "area", "quantity", "subtotal"]
	for d in details:
		for k in NUM_KEYS:
			if d.get(k) in (None, ""):
				d[k] = 0.0
		# total.total
		total_obj = d.get("total")
		if isinstance(total_obj, dict):
			if total_obj.get("total") in (None, ""):
				total_obj["total"] = 0.0
		# discounts
		discounts = d.get("discounts")
		if discounts:
			for disc in discounts:
				for k in NUM_KEYS:
					if disc.get(k) in (None, ""):
						dic_val = 0.0
						disc[k] = dic_val


def extraer_datos_hoja(
	ws,
	*,
	letra_celda_codigo: str = "B",
	numero_celda_codigo: int = 9,
	letra_celda_inicio_elementos: str = "F",
	numero_celda_inicio_elementos: int = 12,
	letra_celda_fin_elementos: str = "S",
	numero_celda_fin_elementos: int = 27,
	steps: int = 33,
) -> List[Dict[str, Any]]:
	"""
	Recorre la hoja para extraer subcategorías con sus quantity_details.

	Devuelve una lista de dicts con forma:
	{
	  "codigo": "14",            # derivado de id (parte antes del punto)
	  "id": "14.1",              # valor leído en B9 (+ n*steps)
	  "total_quantity": 123.0,     # suma de S[row] en el bloque
	  "quantity_details": [        # filas de F..S
		 { location, height, width, length, area, quantity, subtotal,
		   total: { total },
		   discounts: [ { element, height, width, length, area, quantity, subtotal } ]
		 }, ...
	  ]
	}
	"""

	cod_col = letra_celda_codigo.upper()
	elem_c1 = letra_celda_inicio_elementos.upper()
	elem_c2 = letra_celda_fin_elementos.upper()

	# índices de columnas para validación de rango (no imprescindibles, pero útiles si iteramos)
	column_index_from_string(elem_c1)
	column_index_from_string(elem_c2)

	def _normalize(s: Any) -> str:
		return "" if s is None else str(s).strip().upper()

	def _is_codigo_label(val: Any) -> bool:
		lbl = _normalize(val)
		return lbl in ("CODIGO", "CÓDIGO")

	# 1) Detectar encabezados "LOCALIZACION Y/O ELEMENTO" y su columna
	header_cells: List[tuple[int, int]] = []  # (row, col)
	header_text = "LOCALIZACION Y/O ELEMENTO"
	max_row = ws.max_row or 0
	max_col = ws.max_column or 0
	for r in range(1, max_row + 1):
		for c in range(1, max_col + 1):
			v = ws.cell(row=r, column=c).value
			if _normalize(v) == header_text:
				header_cells.append((r, c))
				break  # asumir una coincidencia por fila es suficiente

	subcategorias: List[Dict[str, Any]] = []

	def _row_is_empty(row: int) -> bool:
		# Considerar vacío si todas las celdas F..S están vacías o string vacío
		for col in ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]:
			val = ws[f"{col}{row}"].value
			if val not in (None, ""):
				if not (isinstance(val, str) and val.strip() == ""):
					return False
		return True

	if header_cells:
		header_cells.sort(key=lambda t: (t[0], t[1]))
		for idx, (hr, hc) in enumerate(header_cells):
			# Debajo del título suele haber una fila de encabezados (ALTO, ANCHO, ...), la saltamos
			start_row = hr + 2
			next_hr = header_cells[idx + 1][0] if (idx + 1 < len(header_cells)) else None
			# Límite máximo de 16 filas de elementos
			hard_end = start_row + 16 - 1
			end_row = hard_end
			if next_hr:
				end_row = min(end_row, next_hr - 1)
			end_row = min(end_row, max_row)

			# Ajustar end_row para no pasarnos por filas completamente vacías al final
			def _row_is_empty_rel(row: int, base_col: int) -> bool:
				# Considerar vacío si todas las celdas desde base_col hasta base_col+13 están vacías
				for offset in range(0, 14):  # columnas F..S relativo al encabezado
					col_idx = base_col + offset
					val = ws.cell(row=row, column=col_idx).value
					if val not in (None, ""):
						if not (isinstance(val, str) and str(val).strip() == ""):
							return False
				return True

			r = end_row
			while r >= start_row and _row_is_empty_rel(r, hc):
				r -= 1
			end_row = r

			if end_row < start_row:
				continue  # no hay detalles

			# Buscar id priorizando columna B hacia arriba con preferencia por valores con dígitos,
			# y validando que arriba del id esté la etiqueta 'CODIGO'/'CÓDIGO'.
			def _looks_like_code(val: Any) -> bool:
				if val is None:
					return False
				s = str(val)
				return bool(re.search(r"\d", s))

			id_val = None
			id_pos = None  # (row, col)
			# 1) Columna B (2)
			probe_row = hr - 1
			while probe_row >= 1 and (hr - probe_row) <= 30:
				candidate = ws.cell(row=probe_row, column=2).value
				if candidate not in (None, "") and not (isinstance(candidate, str) and str(candidate).strip() == ""):
					if _looks_like_code(candidate) and _is_codigo_label(ws.cell(row=probe_row - 1, column=2).value):
						id_val = candidate
						id_pos = (probe_row, 2)
						break
				probe_row -= 1

			# 2) Respaldo: columna (hc-2)
			if id_val is None:
				id_col = max(1, hc - 2)
				probe_row = hr - 1
				while probe_row >= 1 and (hr - probe_row) <= 30:
					candidate = ws.cell(row=probe_row, column=id_col).value
					if candidate not in (None, "") and not (isinstance(candidate, str) and str(candidate).strip() == ""):
						if _looks_like_code(candidate) and _is_codigo_label(ws.cell(row=probe_row - 1, column=id_col).value):
							id_val = candidate
							id_pos = (probe_row, id_col)
							break
					probe_row -= 1

			# Si no hay id válido (con dígitos y etiqueta arriba), saltar este bloque
			if id_val is None:
				continue

			details = []
			totals_collected: List[Any] = []
			for row in range(start_row, end_row + 1):
				# Omitir filas completamente vacías en el medio
				if _row_is_empty_rel(row, hc):
					continue

				# Columnas relativas al encabezado: base = hc
				base = hc
				location = ws.cell(row=row, column=base + 0).value
				# Filtrar: incluir solo si location tiene valor
				if location is None or (isinstance(location, str) and str(location).strip() == ""):
					continue

				height = _round2_if_number(ws.cell(row=row, column=base + 1).value)
				width = _round2_if_number(ws.cell(row=row, column=base + 2).value)
				length = _round2_if_number(ws.cell(row=row, column=base + 3).value)
				area = _round2_if_number(ws.cell(row=row, column=base + 4).value)
				quantity = _round2_if_number(ws.cell(row=row, column=base + 5).value)
				subtotal = _round2_if_number(ws.cell(row=row, column=base + 6).value)

				d_height = _round2_if_number(ws.cell(row=row, column=base + 7).value)
				d_width = _round2_if_number(ws.cell(row=row, column=base + 8).value)
				d_length = _round2_if_number(ws.cell(row=row, column=base + 9).value)
				d_area = _round2_if_number(ws.cell(row=row, column=base + 10).value)
				d_quantity = _round2_if_number(ws.cell(row=row, column=base + 11).value)
				d_subtotal = _round2_if_number(ws.cell(row=row, column=base + 12).value)

				total_val = _round2_if_number(ws.cell(row=row, column=base + 13).value)
				totals_collected.append(total_val)

				details.append(
					{
						"location": location,
						"height": height,
						"width": width,
						"length": length,
						"area": area,
						"quantity": quantity,
						"subtotal": subtotal,
						"total": {"total": total_val},
						"discounts": [
							{
								"element": "",
								"height": d_height,
								"width": d_width,
								"length": d_length,
								"area": d_area,
								"quantity": d_quantity,
								"subtotal": d_subtotal,
							}
						],
					}
				)

			# Limpiar descuentos vacíos y armar subcategoría; omitir total_quantity si no hay código numérico
			codigo_cat = _categoria_from_id(id_val)
			has_numeric_code = any(ch.isdigit() for ch in str(id_val))
			_clean_discounts_in_details(details)
			_zero_nulls_in_details(details)
			subcat = {
				"codigo": codigo_cat,
				"id": id_val,
				"quantity_details": details,
			}
			if has_numeric_code:
				subcat["total_quantity"] = _round2_if_number(_sum_safe(totals_collected))
			# Agregar solo si hay detalles
			if details:
				subcategorias.append(subcat)

		return subcategorias

	# 2) Fallback: método por pasos (si no se halló el encabezado)
	subcategorias = []
	i = 0
	while True:
		code_row = numero_celda_codigo + i * steps
		code_cell = f"{cod_col}{code_row}"
		id_val = ws[code_cell].value

		if id_val is None or (isinstance(id_val, str) and id_val.strip() == ""):
			break

		# Validar que sobre el ID esté la etiqueta CODIGO/CÓDIGO en la misma columna del código
		label_row = code_row - 1
		if label_row >= 1:
			if not _is_codigo_label(ws[f"{cod_col}{label_row}"].value):
				i += 1
				continue

		r1 = numero_celda_inicio_elementos + i * steps
		r2 = numero_celda_fin_elementos + i * steps

		details = []
		totals_collected = []
		for row in range(r1, r2 + 1):
			location = ws[f"F{row}"].value
			if location is None or (isinstance(location, str) and location.strip() == ""):
				continue
			height = _round2_if_number(ws[f"G{row}"].value)
			width = _round2_if_number(ws[f"H{row}"].value)
			length = _round2_if_number(ws[f"I{row}"].value)
			area = _round2_if_number(ws[f"J{row}"].value)
			quantity = _round2_if_number(ws[f"K{row}"].value)
			subtotal = _round2_if_number(ws[f"L{row}"].value)

			d_height = _round2_if_number(ws[f"M{row}"].value)
			d_width = _round2_if_number(ws[f"N{row}"].value)
			d_length = _round2_if_number(ws[f"O{row}"].value)
			d_area = _round2_if_number(ws[f"P{row}"].value)
			d_quantity = _round2_if_number(ws[f"Q{row}"].value)
			d_subtotal = _round2_if_number(ws[f"R{row}"].value)

			total_val = _round2_if_number(ws[f"S{row}"].value)
			totals_collected.append(total_val)

			details.append(
				{
					"location": location,
					"height": height,
					"width": width,
					"length": length,
					"area": area,
					"quantity": quantity,
					"subtotal": subtotal,
					"total": {"total": total_val},
					"discounts": [
						{
							"element": "",
							"height": d_height,
							"width": d_width,
							"length": d_length,
							"area": d_area,
							"quantity": d_quantity,
							"subtotal": d_subtotal,
						}
					],
				}
			)

		codigo_cat = _categoria_from_id(id_val)
		has_numeric_code = any(ch.isdigit() for ch in str(id_val))
		_clean_discounts_in_details(details)
		_zero_nulls_in_details(details)
		subcat = {
			"codigo": codigo_cat,
			"id": id_val,
			"quantity_details": details,
		}
		if has_numeric_code:
			subcat["total_quantity"] = _round2_if_number(_sum_safe(totals_collected))
		# Agregar solo si hay detalles
		if details:
			subcategorias.append(subcat)

		i += 1

	return subcategorias


def procesar_archivo(
	xlsx_path: Path,
	*,
	letra_celda_codigo: str,
	numero_celda_codigo: int,
	letra_celda_inicio_elementos: str,
	numero_celda_inicio_elementos: int,
	letra_celda_fin_elementos: str,
	numero_celda_fin_elementos: int,
	steps: int,
) -> Optional[Dict[str, Any]]:
	"""Procesa un archivo .xlsx y devuelve un dict listo para exportar a JSON.

	Estructura devuelta:
	{
	  "archivo": nombre_base.xlsx,
	  "cedula": valor_de_APU_L6,
	  "datos": [ {"codigo": ..., "elementos": [[...], ...]}, ... ]
	}
	"""
	try:
		wb = load_workbook(filename=str(xlsx_path), data_only=True)
	except Exception as exc:
		print(f"[ERROR] No se pudo abrir '{xlsx_path.name}': {exc}")
		return None

	if "APU" not in wb.sheetnames:
		print(f"[ADVERTENCIA] Hoja 'APU' no encontrada en '{xlsx_path.name}'. Se omite.")
		return None

	apu_index = wb.sheetnames.index("APU")
	apu_ws = wb["APU"]
	cedula = leer_cedula(apu_ws)

	# Hoja objetivo: preferir 'CANT. BENEFICIARIO 1' si existe; si no, usar la siguiente a APU
	next_index = apu_index + 1
	if next_index >= len(wb.sheetnames):
		print(f"[ADVERTENCIA] No existe hoja siguiente a 'APU' en '{xlsx_path.name}'. Se omite.")
		return None

	desired_name = "CANT. BENEFICIARIO 1"
	next_sheet_name = wb.sheetnames[next_index]
	# Si la hoja inmediata no es la deseada, pero existe la deseada en el libro, usarla
	if next_sheet_name != desired_name and desired_name in wb.sheetnames:
		target_sheet_name = desired_name
	else:
		target_sheet_name = next_sheet_name

	ws_target = wb[target_sheet_name]

	subcats = extraer_datos_hoja(
		ws_target,
		letra_celda_codigo=letra_celda_codigo,
		numero_celda_codigo=numero_celda_codigo,
		letra_celda_inicio_elementos=letra_celda_inicio_elementos,
		numero_celda_inicio_elementos=numero_celda_inicio_elementos,
		letra_celda_fin_elementos=letra_celda_fin_elementos,
		numero_celda_fin_elementos=numero_celda_fin_elementos,
		steps=steps,
	)

	# Agrupar por categoría (codigo)
	cats_map: Dict[str, Dict[str, Any]] = {}
	for sc in subcats:
		codigo_cat = _to_str(sc.get("codigo", ""))
		if codigo_cat not in cats_map:
			cats_map[codigo_cat] = {"codigo": codigo_cat, "subcategories": []}
		# Subcategorías incluyen id, total_quantity y quantity_details
		sub = {
			"id": sc.get("id"),
			"quantity_details": sc.get("quantity_details", []),
		}
		if "total_quantity" in sc:
			sub["total_quantity"] = sc["total_quantity"]
		cats_map[codigo_cat]["subcategories"].append(sub)

	categories = list(cats_map.values())

	return {
		"cedula": cedula,
		"categories": categories,
	}


def siguiente_consecutivo(output_dir: Path) -> int:
	"""Calcula el siguiente consecutivo basado en archivos N.json existentes."""
	max_n = 0
	for p in output_dir.glob("*.json"):
		name = p.stem
		if name.isdigit():
			n = int(name)
			if n > max_n:
				max_n = n
	return max_n + 1


def guardar_json_con_consecutivo(output_dir: Path, data: Dict[str, Any]) -> Path:
	output_dir.mkdir(parents=True, exist_ok=True)
	n = siguiente_consecutivo(output_dir)
	out_path = output_dir / f"{n}.json"
	with out_path.open("w", encoding="utf-8") as f:
		json.dump(data, f, ensure_ascii=False, indent=2)
	return out_path


def listar_excels(input_dir: Path) -> List[Path]:
	"""Lista archivos .xlsx válidos (excluye temporales ~) en input_dir."""
	if not input_dir.exists():
		input_dir.mkdir(parents=True, exist_ok=True)
	return [p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~$")]


def parse_args():
	parser = argparse.ArgumentParser(
		description=(
			"Extrae cédula de APU!L6 y pares {código, elementos} de la hoja siguiente a APU, "
			"para múltiples .xlsx y exporta JSON consecutivos."
		)
	)
	parser.add_argument(
		"--input-dir",
		default="input_xlsx",
		help="Carpeta con los .xlsx a procesar (por defecto: input_xlsx)",
	)
	parser.add_argument(
		"--output-dir",
		default="output_json",
		help="Carpeta donde se guardarán los JSON (por defecto: output_json)",
	)
	parser.add_argument("--steps", type=int, default=33, help="Número de filas entre bloques (por defecto: 33)")
	parser.add_argument("--code-col", default="B", help="Columna del código (por defecto: B)")
	parser.add_argument("--code-row-start", type=int, default=9, help="Fila inicial del código (por defecto: 9)")
	parser.add_argument("--elem-col-start", default="F", help="Columna inicial del rango de elementos (por defecto: F)")
	parser.add_argument("--elem-row-start", type=int, default=12, help="Fila inicial del rango de elementos (por defecto: 12)")
	parser.add_argument("--elem-col-end", default="S", help="Columna final del rango de elementos (por defecto: S)")
	parser.add_argument("--elem-row-end", type=int, default=27, help="Fila final del rango de elementos (por defecto: 27)")
	return parser.parse_args()


def main():
	args = parse_args()

	input_dir = Path(args.input_dir)
	output_dir = Path(args.output_dir)

	excels = listar_excels(input_dir)
	if not excels:
		print(
			f"[INFO] No se encontraron .xlsx en '{input_dir.resolve()}'. "
			f"Coloca tus archivos allí y vuelve a ejecutar."
		)
		return

	print(f"[INFO] Archivos a procesar: {len(excels)} en '{input_dir.resolve()}'")

	procesados = 0
	for xlsx in sorted(excels):
		print(f"[INFO] Procesando: {xlsx.name}")

		data = procesar_archivo(
			xlsx,
			letra_celda_codigo=args.code_col,
			numero_celda_codigo=args.code_row_start,
			letra_celda_inicio_elementos=args.elem_col_start,
			numero_celda_inicio_elementos=args.elem_row_start,
			letra_celda_fin_elementos=args.elem_col_end,
			numero_celda_fin_elementos=args.elem_row_end,
			steps=args.steps,
		)

		if data is None:
			continue

		out_path = guardar_json_con_consecutivo(output_dir, data)
		print(f"[OK] Exportado: {out_path}")
		procesados += 1

	print(f"[RESUMEN] Exportados {procesados} JSON en '{output_dir.resolve()}'")


if __name__ == "__main__":
	main()

